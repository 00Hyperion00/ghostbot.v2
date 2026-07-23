from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import secrets
import shutil
import sqlite3
import tempfile
import zipfile
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from functools import wraps
from pathlib import Path
from typing import Any, Iterable

from flask import (
    Blueprint,
    Response,
    abort,
    current_app,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

VERSION = "3.0.0"
BP = Blueprint(
    "enterprise_v3",
    __name__,
    url_prefix="/v3",
    template_folder="templates",
    static_folder="static",
    static_url_path="/assets",
)

ROLE_SEED = {
    "SYS_ADMIN": ("Sistem Yöneticisi", ["*"]),
    "AUDITOR": ("SMMM / Denetçi", ["dashboard.view", "company.view", "audit.view", "task.manage", "aging.manage", "bank.manage", "report.manage", "backup.create", "log.view"]),
    "ASSISTANT": ("Denetçi Yardımcısı", ["dashboard.view", "company.view", "audit.view", "task.manage", "aging.manage", "bank.manage", "report.view"]),
    "ACCOUNTING": ("Muhasebe Sorumlusu", ["dashboard.view", "company.view", "audit.view", "task.update", "aging.view", "bank.view", "report.view"]),
    "CLIENT": ("Firma Kullanıcısı", ["dashboard.view", "company.view", "task.update", "report.view"]),
    "EXECUTIVE": ("Salt Okunur Yönetici", ["dashboard.view", "company.view", "audit.view", "aging.view", "bank.view", "report.view"]),
}

PERMISSIONS = {
    "dashboard.view": "Kurumsal gösterge paneli",
    "company.view": "Firma verilerini görüntüleme",
    "company.manage": "Firma ve grup yönetimi",
    "audit.view": "Denetim ve bulguları görüntüleme",
    "task.manage": "Görev oluşturma ve güncelleme",
    "task.update": "Atanmış görevi güncelleme",
    "aging.manage": "Cari yaşlandırma yükleme/yönetme",
    "aging.view": "Cari yaşlandırma görüntüleme",
    "bank.manage": "Banka hareketi yükleme ve eşleştirme",
    "bank.view": "Banka eşleştirme görüntüleme",
    "report.manage": "Rapor şablonu ve çıktı yönetimi",
    "report.view": "Rapor görüntüleme",
    "user.manage": "Kullanıcı ve rol yönetimi",
    "backup.create": "Yedek oluşturma ve indirme",
    "backup.restore": "Yedek geri yükleme",
    "log.view": "Denetim ve güvenlik günlüğü",
    "company.all": "Tüm firmalara erişim",
}

SCHEMA = """
CREATE TABLE IF NOT EXISTS v3_users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT NOT NULL UNIQUE,
    full_name TEXT NOT NULL,
    email TEXT,
    password_hash TEXT NOT NULL,
    active INTEGER NOT NULL DEFAULT 1,
    must_change_password INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    last_login TEXT
);
CREATE TABLE IF NOT EXISTS v3_roles (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    code TEXT NOT NULL UNIQUE,
    name TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS v3_permissions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    code TEXT NOT NULL UNIQUE,
    name TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS v3_role_permissions (
    role_id INTEGER NOT NULL,
    permission_id INTEGER NOT NULL,
    PRIMARY KEY(role_id, permission_id)
);
CREATE TABLE IF NOT EXISTS v3_companies (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    tax_no TEXT,
    group_id INTEGER,
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS v3_company_groups (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    active INTEGER NOT NULL DEFAULT 1
);
CREATE TABLE IF NOT EXISTS v3_user_roles (
    user_id INTEGER NOT NULL,
    role_id INTEGER NOT NULL,
    company_id INTEGER,
    PRIMARY KEY(user_id, role_id, company_id)
);
CREATE TABLE IF NOT EXISTS v3_tasks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    company_id INTEGER,
    audit_id INTEGER,
    finding_id INTEGER,
    title TEXT NOT NULL,
    description TEXT,
    assigned_to INTEGER,
    due_date TEXT,
    status TEXT NOT NULL DEFAULT 'Yeni',
    priority TEXT NOT NULL DEFAULT 'Orta',
    created_by INTEGER,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    completed_at TEXT
);
CREATE TABLE IF NOT EXISTS v3_notifications (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    title TEXT NOT NULL,
    body TEXT,
    target_url TEXT,
    is_read INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS v3_aging_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    company_id INTEGER NOT NULL,
    audit_id INTEGER,
    account_code TEXT,
    counterparty_code TEXT,
    counterparty_name TEXT,
    tax_no TEXT,
    document_no TEXT,
    document_date TEXT,
    due_date TEXT,
    debit REAL NOT NULL DEFAULT 0,
    credit REAL NOT NULL DEFAULT 0,
    balance REAL NOT NULL DEFAULT 0,
    age_days INTEGER NOT NULL DEFAULT 0,
    age_bucket TEXT,
    source_hash TEXT,
    created_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_v3_aging_company ON v3_aging_items(company_id, audit_id);
CREATE TABLE IF NOT EXISTS v3_bank_imports (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    company_id INTEGER NOT NULL,
    audit_id INTEGER,
    filename TEXT NOT NULL,
    sha256 TEXT NOT NULL,
    row_count INTEGER NOT NULL DEFAULT 0,
    created_by INTEGER,
    created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS v3_bank_transactions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    import_id INTEGER NOT NULL,
    company_id INTEGER NOT NULL,
    audit_id INTEGER,
    account_no TEXT,
    txn_date TEXT,
    value_date TEXT,
    reference TEXT,
    description TEXT,
    currency TEXT,
    amount REAL NOT NULL DEFAULT 0,
    balance REAL,
    match_status TEXT NOT NULL DEFAULT 'Eşleşmedi',
    match_score INTEGER NOT NULL DEFAULT 0,
    journal_table TEXT,
    journal_row_id TEXT,
    journal_reference TEXT,
    journal_date TEXT,
    journal_amount REAL,
    created_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_v3_bank_company ON v3_bank_transactions(company_id, audit_id, match_status);
CREATE TABLE IF NOT EXISTS v3_report_templates (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    company_id INTEGER,
    report_type TEXT NOT NULL DEFAULT 'Yönetim',
    config_json TEXT NOT NULL,
    active INTEGER NOT NULL DEFAULT 1,
    created_by INTEGER,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS v3_backup_jobs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    filename TEXT NOT NULL,
    sha256 TEXT NOT NULL,
    encrypted INTEGER NOT NULL DEFAULT 0,
    size_bytes INTEGER NOT NULL,
    created_by INTEGER,
    created_at TEXT NOT NULL,
    verified_at TEXT
);
CREATE TABLE IF NOT EXISTS v3_audit_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER,
    username TEXT,
    action TEXT NOT NULL,
    entity_type TEXT,
    entity_id TEXT,
    company_id INTEGER,
    old_value TEXT,
    new_value TEXT,
    ip_address TEXT,
    user_agent TEXT,
    created_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_v3_log_time ON v3_audit_log(created_at DESC);
CREATE TABLE IF NOT EXISTS v3_settings (
    key TEXT PRIMARY KEY,
    value TEXT,
    updated_at TEXT NOT NULL
);
"""


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat(sep=" ")


def db_path() -> Path:
    configured = current_app.config.get("V3_DATABASE") or current_app.config.get("DATABASE")
    if configured:
        return Path(configured)
    return Path(current_app.root_path) / "data" / "audit.db"


def connect() -> sqlite3.Connection:
    path = db_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=5000")
    return conn


def init_enterprise_db(app=None) -> None:
    if app is not None:
        with app.app_context():
            return init_enterprise_db()
    with connect() as conn:
        conn.executescript(SCHEMA)
        for code, name in PERMISSIONS.items():
            conn.execute("INSERT OR IGNORE INTO v3_permissions(code,name) VALUES(?,?)", (code, name))
        for role_code, (role_name, permissions) in ROLE_SEED.items():
            conn.execute("INSERT OR IGNORE INTO v3_roles(code,name) VALUES(?,?)", (role_code, role_name))
            role_id = conn.execute("SELECT id FROM v3_roles WHERE code=?", (role_code,)).fetchone()[0]
            if permissions == ["*"]:
                permission_rows = conn.execute("SELECT id FROM v3_permissions").fetchall()
            else:
                q = ",".join("?" for _ in permissions)
                permission_rows = conn.execute(f"SELECT id FROM v3_permissions WHERE code IN ({q})", permissions).fetchall()
            for row in permission_rows:
                conn.execute("INSERT OR IGNORE INTO v3_role_permissions(role_id,permission_id) VALUES(?,?)", (role_id, row[0]))
        conn.commit()


def ensure_secret(app) -> None:
    if app.secret_key:
        return
    secret_file = Path(app.root_path) / "data" / "v3_secret.key"
    secret_file.parent.mkdir(parents=True, exist_ok=True)
    if secret_file.exists():
        app.secret_key = secret_file.read_text(encoding="utf-8").strip()
    else:
        app.secret_key = secrets.token_urlsafe(48)
        secret_file.write_text(app.secret_key, encoding="utf-8")


def create_or_update_admin(username: str, password: str, full_name: str = "Sistem Yöneticisi") -> None:
    with connect() as conn:
        ts = now_iso()
        existing = conn.execute("SELECT id FROM v3_users WHERE username=?", (username,)).fetchone()
        if existing:
            user_id = existing[0]
            conn.execute("UPDATE v3_users SET password_hash=?, active=1, must_change_password=1, full_name=? WHERE id=?", (generate_password_hash(password), full_name, user_id))
        else:
            cur = conn.execute("INSERT INTO v3_users(username,full_name,password_hash,active,must_change_password,created_at) VALUES(?,?,?,?,?,?)", (username, full_name, generate_password_hash(password), 1, 1, ts))
            user_id = cur.lastrowid
        role_id = conn.execute("SELECT id FROM v3_roles WHERE code='SYS_ADMIN'").fetchone()[0]
        conn.execute("INSERT OR IGNORE INTO v3_user_roles(user_id,role_id,company_id) VALUES(?,?,NULL)", (user_id, role_id))
        conn.commit()


def register_enterprise_v3(app) -> None:
    if "enterprise_v3" in app.blueprints:
        return
    ensure_secret(app)
    app.register_blueprint(BP)
    init_enterprise_db(app)

    @app.context_processor
    def _v3_context():
        return {"v3_version": VERSION, "v3_user": getattr(g, "v3_user", None)}


def table_exists(conn: sqlite3.Connection, table: str) -> bool:
    return conn.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone() is not None


def table_columns(conn: sqlite3.Connection, table: str) -> list[str]:
    return [row[1] for row in conn.execute(f'PRAGMA table_info("{table}")').fetchall()]


def csrf_token() -> str:
    if "v3_csrf" not in session:
        session["v3_csrf"] = secrets.token_urlsafe(32)
    return session["v3_csrf"]


def audit(action: str, entity_type: str | None = None, entity_id: Any = None, company_id: Any = None, old: Any = None, new: Any = None) -> None:
    user = getattr(g, "v3_user", None)
    with connect() as conn:
        conn.execute(
            "INSERT INTO v3_audit_log(user_id,username,action,entity_type,entity_id,company_id,old_value,new_value,ip_address,user_agent,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
            (
                user["id"] if user else None,
                user["username"] if user else None,
                action,
                entity_type,
                str(entity_id) if entity_id is not None else None,
                company_id,
                json.dumps(old, ensure_ascii=False, default=str) if old is not None else None,
                json.dumps(new, ensure_ascii=False, default=str) if new is not None else None,
                request.headers.get("X-Forwarded-For", request.remote_addr),
                request.headers.get("User-Agent", "")[:500],
                now_iso(),
            ),
        )
        conn.commit()


def current_permissions() -> set[str]:
    user = getattr(g, "v3_user", None)
    if not user:
        return set()
    with connect() as conn:
        rows = conn.execute(
            "SELECT DISTINCT p.code FROM v3_permissions p JOIN v3_role_permissions rp ON rp.permission_id=p.id JOIN v3_user_roles ur ON ur.role_id=rp.role_id WHERE ur.user_id=?",
            (user["id"],),
        ).fetchall()
    return {r[0] for r in rows}


def has_permission(code: str) -> bool:
    return code in current_permissions()


def permission_required(code: str):
    def decorator(fn):
        @wraps(fn)
        def wrapped(*args, **kwargs):
            if not has_permission(code):
                abort(403)
            return fn(*args, **kwargs)
        return wrapped
    return decorator


def accessible_company_ids() -> list[int]:
    user = getattr(g, "v3_user", None)
    if not user:
        return []
    with connect() as conn:
        if has_permission("company.all"):
            return [r[0] for r in conn.execute("SELECT id FROM v3_companies WHERE active=1").fetchall()]
        return [r[0] for r in conn.execute("SELECT DISTINCT company_id FROM v3_user_roles WHERE user_id=? AND company_id IS NOT NULL", (user["id"],)).fetchall()]


def company_scope_sql(alias: str = "company_id") -> tuple[str, list[Any]]:
    ids = accessible_company_ids()
    if has_permission("company.all"):
        return "1=1", []
    if not ids:
        return "1=0", []
    return f"{alias} IN ({','.join('?' for _ in ids)})", ids


def parse_decimal(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    text = str(value).strip().replace("₺", "").replace("TL", "").replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError):
        return 0.0


def parse_date(value: Any) -> str | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return None


def read_tabular(file_storage) -> list[dict[str, Any]]:
    filename = secure_filename(file_storage.filename or "")
    raw = file_storage.read()
    suffix = Path(filename).suffix.lower()
    if suffix in {".csv", ".txt"}:
        text = None
        for enc in ("utf-8-sig", "utf-16", "cp1254", "latin1"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                pass
        if text is None:
            raise ValueError("Dosya kodlaması okunamadı")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        return [dict(row) for row in csv.DictReader(io.StringIO(text), dialect=dialect)]
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(v).strip() if v is not None else "" for v in next(rows)]
        return [{headers[i]: row[i] for i in range(min(len(headers), len(row)))} for row in rows if any(v is not None for v in row)]
    raise ValueError("CSV, TXT veya XLSX yükleyin")


def pick(row: dict[str, Any], *names: str) -> Any:
    normalized = {str(k).strip().lower().replace("ı", "i"): v for k, v in row.items()}
    for name in names:
        key = name.lower().replace("ı", "i")
        if key in normalized:
            return normalized[key]
    for key, value in normalized.items():
        if any(name.lower().replace("ı", "i") in key for name in names):
            return value
    return None


def age_bucket(days: int) -> str:
    if days <= 30: return "0-30"
    if days <= 60: return "31-60"
    if days <= 90: return "61-90"
    if days <= 180: return "91-180"
    if days <= 365: return "181-365"
    return "365+"


def paginate(page: int, per_page: int, total: int) -> dict[str, int]:
    pages = max(1, (total + per_page - 1) // per_page)
    return {"page": page, "per_page": per_page, "total": total, "pages": pages}


@BP.before_request
def before_request():
    g.v3_user = None
    user_id = session.get("v3_user_id")
    if user_id:
        with connect() as conn:
            g.v3_user = conn.execute("SELECT * FROM v3_users WHERE id=? AND active=1", (user_id,)).fetchone()
    if request.method == "POST":
        supplied = request.form.get("csrf_token") or request.headers.get("X-CSRF-Token")
        if not supplied or not secrets.compare_digest(supplied, session.get("v3_csrf", "")):
            abort(400, "Geçersiz CSRF doğrulaması")
    exempt = {"enterprise_v3.login", "enterprise_v3.static", "enterprise_v3.health"}
    if request.endpoint not in exempt and g.v3_user is None:
        return redirect(url_for("enterprise_v3.login", next=request.full_path))


@BP.app_context_processor
def inject_helpers():
    unread = 0
    if getattr(g, "v3_user", None):
        with connect() as conn:
            unread = conn.execute("SELECT COUNT(*) FROM v3_notifications WHERE user_id=? AND is_read=0", (g.v3_user["id"],)).fetchone()[0]
    return {"csrf_token": csrf_token, "v3_has_permission": has_permission, "v3_unread_notifications": unread}


@BP.route("/health")
def health():
    return jsonify({"ok": True, "version": VERSION})


@BP.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        with connect() as conn:
            user = conn.execute("SELECT * FROM v3_users WHERE username=? AND active=1", (username,)).fetchone()
            if user and check_password_hash(user["password_hash"], password):
                session.clear()
                session["v3_user_id"] = user["id"]
                session["v3_csrf"] = secrets.token_urlsafe(32)
                conn.execute("UPDATE v3_users SET last_login=? WHERE id=?", (now_iso(), user["id"]))
                conn.commit()
                g.v3_user = user
                audit("LOGIN", "user", user["id"])
                return redirect(request.args.get("next") or url_for("enterprise_v3.dashboard"))
        flash("Kullanıcı adı veya parola hatalı.", "danger")
    return render_template("v3/login.html")


@BP.route("/logout")
def logout():
    if g.v3_user:
        audit("LOGOUT", "user", g.v3_user["id"])
    session.clear()
    return redirect(url_for("enterprise_v3.login"))


@BP.route("/change-password", methods=["POST"])
def change_password():
    old = request.form.get("old_password", "")
    new = request.form.get("new_password", "")
    if len(new) < 10:
        flash("Yeni parola en az 10 karakter olmalıdır.", "danger")
        return redirect(url_for("enterprise_v3.dashboard"))
    if not check_password_hash(g.v3_user["password_hash"], old):
        flash("Mevcut parola hatalı.", "danger")
        return redirect(url_for("enterprise_v3.dashboard"))
    with connect() as conn:
        conn.execute("UPDATE v3_users SET password_hash=?, must_change_password=0 WHERE id=?", (generate_password_hash(new), g.v3_user["id"]))
        conn.commit()
    audit("PASSWORD_CHANGED", "user", g.v3_user["id"])
    flash("Parola güncellendi.", "success")
    return redirect(url_for("enterprise_v3.dashboard"))


def existing_finding_metrics(conn: sqlite3.Connection, company_ids: list[int]) -> dict[str, Any]:
    result = {"open_findings": 0, "critical_findings": 0, "financial_risk": 0.0}
    candidates = ["findings", "audit_findings", "bulgular"]
    table = next((t for t in candidates if table_exists(conn, t)), None)
    if not table:
        return result
    cols = table_columns(conn, table)
    status_col = next((c for c in cols if c.lower() in {"status", "durum"}), None)
    risk_col = next((c for c in cols if c.lower() in {"risk_level", "severity", "risk", "risk_seviyesi"}), None)
    amount_col = next((c for c in cols if c.lower() in {"amount", "tutar", "financial_effect", "parasal_etki"}), None)
    company_col = next((c for c in cols if c.lower() in {"company_id", "firma_id"}), None)
    where, params = [], []
    if company_col and company_ids and not has_permission("company.all"):
        where.append(f'"{company_col}" IN ({",".join("?" for _ in company_ids)})')
        params.extend(company_ids)
    sql_where = " WHERE " + " AND ".join(where) if where else ""
    rows = conn.execute(f'SELECT * FROM "{table}"{sql_where}', params).fetchall()
    for row in rows:
        status = str(row[status_col] if status_col else "").lower()
        if status not in {"kapatıldı", "closed", "düzeltildi", "yanlış pozitif"}:
            result["open_findings"] += 1
        risk = str(row[risk_col] if risk_col else "").lower()
        if risk in {"kritik", "critical"} and status not in {"kapatıldı", "closed"}:
            result["critical_findings"] += 1
        if amount_col:
            result["financial_risk"] += abs(parse_decimal(row[amount_col]))
    return result


@BP.route("/")
@permission_required("dashboard.view")
def dashboard():
    company_ids = accessible_company_ids()
    with connect() as conn:
        scope, params = company_scope_sql("company_id")
        task = conn.execute(f"SELECT COUNT(*) total, SUM(CASE WHEN status NOT IN ('Tamamlandı','İptal') THEN 1 ELSE 0 END) open, SUM(CASE WHEN status NOT IN ('Tamamlandı','İptal') AND due_date < date('now') THEN 1 ELSE 0 END) overdue FROM v3_tasks WHERE {scope}", params).fetchone()
        aging = conn.execute(f"SELECT COALESCE(SUM(balance),0) total, COALESCE(SUM(CASE WHEN age_days>90 THEN balance ELSE 0 END),0) over90 FROM v3_aging_items WHERE {scope}", params).fetchone()
        bank = conn.execute(f"SELECT COUNT(*) total, SUM(CASE WHEN match_status='Eşleşti' THEN 1 ELSE 0 END) matched FROM v3_bank_transactions WHERE {scope}", params).fetchone()
        companies = conn.execute("SELECT * FROM v3_companies WHERE active=1 ORDER BY name").fetchall() if has_permission("company.all") else conn.execute(f"SELECT * FROM v3_companies WHERE id IN ({','.join('?' for _ in company_ids)}) ORDER BY name", company_ids).fetchall() if company_ids else []
        findings = existing_finding_metrics(conn, company_ids)
        recent_tasks = conn.execute(f"SELECT t.*, c.name company_name, u.full_name assigned_name FROM v3_tasks t LEFT JOIN v3_companies c ON c.id=t.company_id LEFT JOIN v3_users u ON u.id=t.assigned_to WHERE {scope.replace('company_id','t.company_id')} ORDER BY CASE WHEN t.due_date IS NULL THEN 1 ELSE 0 END, t.due_date LIMIT 10", params).fetchall()
    metrics = {
        "company_count": len(companies),
        "open_tasks": task["open"] or 0,
        "overdue_tasks": task["overdue"] or 0,
        "aging_total": aging["total"] or 0,
        "aging_over90": aging["over90"] or 0,
        "bank_match_rate": round(((bank["matched"] or 0) / bank["total"] * 100), 1) if bank["total"] else 0,
        **findings,
    }
    return render_template("v3/dashboard.html", metrics=metrics, companies=companies, recent_tasks=recent_tasks)


@BP.route("/companies", methods=["GET", "POST"])
@permission_required("company.manage")
def companies():
    with connect() as conn:
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            if not name:
                flash("Firma adı zorunludur.", "danger")
            else:
                cur = conn.execute("INSERT INTO v3_companies(name,tax_no,group_id,active,created_at) VALUES(?,?,?,?,?)", (name, request.form.get("tax_no"), request.form.get("group_id") or None, 1, now_iso()))
                conn.commit()
                audit("CREATE", "company", cur.lastrowid, cur.lastrowid, new={"name": name})
                flash("Firma oluşturuldu.", "success")
        rows = conn.execute("SELECT c.*, g.name group_name FROM v3_companies c LEFT JOIN v3_company_groups g ON g.id=c.group_id ORDER BY c.name").fetchall()
        groups = conn.execute("SELECT * FROM v3_company_groups WHERE active=1 ORDER BY name").fetchall()
    return render_template("v3/companies.html", companies=rows, groups=groups)


@BP.route("/users", methods=["GET", "POST"])
@permission_required("user.manage")
def users():
    with connect() as conn:
        if request.method == "POST":
            action = request.form.get("action", "create")
            if action == "create":
                username = request.form.get("username", "").strip()
                password = request.form.get("password", "")
                full_name = request.form.get("full_name", "").strip()
                if not username or len(password) < 10 or not full_name:
                    flash("Ad, kullanıcı adı ve en az 10 karakterli parola zorunludur.", "danger")
                else:
                    try:
                        cur = conn.execute("INSERT INTO v3_users(username,full_name,email,password_hash,active,must_change_password,created_at) VALUES(?,?,?,?,1,1,?)", (username, full_name, request.form.get("email"), generate_password_hash(password), now_iso()))
                        user_id = cur.lastrowid
                        role_id = int(request.form.get("role_id"))
                        company_id = request.form.get("company_id") or None
                        conn.execute("INSERT INTO v3_user_roles(user_id,role_id,company_id) VALUES(?,?,?)", (user_id, role_id, company_id))
                        conn.commit()
                        audit("CREATE", "user", user_id, company_id, new={"username": username})
                        flash("Kullanıcı oluşturuldu.", "success")
                    except sqlite3.IntegrityError:
                        flash("Bu kullanıcı adı zaten mevcut.", "danger")
            elif action == "toggle":
                user_id = int(request.form.get("user_id"))
                old = conn.execute("SELECT active FROM v3_users WHERE id=?", (user_id,)).fetchone()
                if old and user_id != g.v3_user["id"]:
                    conn.execute("UPDATE v3_users SET active=? WHERE id=?", (0 if old[0] else 1, user_id))
                    conn.commit()
                    audit("TOGGLE_ACTIVE", "user", user_id, old={"active": old[0]}, new={"active": 0 if old[0] else 1})
        rows = conn.execute("SELECT u.*, GROUP_CONCAT(r.name || COALESCE(' / ' || c.name,''), ', ') role_names FROM v3_users u LEFT JOIN v3_user_roles ur ON ur.user_id=u.id LEFT JOIN v3_roles r ON r.id=ur.role_id LEFT JOIN v3_companies c ON c.id=ur.company_id GROUP BY u.id ORDER BY u.full_name").fetchall()
        roles = conn.execute("SELECT * FROM v3_roles ORDER BY name").fetchall()
        company_rows = conn.execute("SELECT * FROM v3_companies WHERE active=1 ORDER BY name").fetchall()
    return render_template("v3/users.html", users=rows, roles=roles, companies=company_rows)


@BP.route("/tasks", methods=["GET", "POST"])
def tasks():
    if request.method == "POST":
        if not (has_permission("task.manage") or has_permission("task.update")):
            abort(403)
        with connect() as conn:
            action = request.form.get("action", "create")
            if action == "create" and has_permission("task.manage"):
                company_id = int(request.form.get("company_id"))
                if company_id not in accessible_company_ids() and not has_permission("company.all"):
                    abort(403)
                cur = conn.execute("INSERT INTO v3_tasks(company_id,audit_id,finding_id,title,description,assigned_to,due_date,status,priority,created_by,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)", (company_id, request.form.get("audit_id") or None, request.form.get("finding_id") or None, request.form.get("title", "").strip(), request.form.get("description"), request.form.get("assigned_to") or None, request.form.get("due_date") or None, "Yeni", request.form.get("priority", "Orta"), g.v3_user["id"], now_iso(), now_iso()))
                task_id = cur.lastrowid
                if request.form.get("assigned_to"):
                    conn.execute("INSERT INTO v3_notifications(user_id,title,body,target_url,created_at) VALUES(?,?,?,?,?)", (request.form.get("assigned_to"), "Yeni görev atandı", request.form.get("title"), url_for("enterprise_v3.tasks"), now_iso()))
                conn.commit()
                audit("CREATE", "task", task_id, company_id)
            else:
                task_id = int(request.form.get("task_id"))
                task = conn.execute("SELECT * FROM v3_tasks WHERE id=?", (task_id,)).fetchone()
                if not task:
                    abort(404)
                if not has_permission("task.manage") and task["assigned_to"] != g.v3_user["id"]:
                    abort(403)
                status = request.form.get("status", task["status"])
                completed = now_iso() if status == "Tamamlandı" else None
                conn.execute("UPDATE v3_tasks SET status=?, priority=?, assigned_to=?, due_date=?, description=?, updated_at=?, completed_at=? WHERE id=?", (status, request.form.get("priority", task["priority"]), request.form.get("assigned_to") or task["assigned_to"], request.form.get("due_date") or task["due_date"], request.form.get("description", task["description"]), now_iso(), completed, task_id))
                conn.commit()
                audit("UPDATE", "task", task_id, task["company_id"], old=dict(task), new={"status": status})
        return redirect(url_for("enterprise_v3.tasks"))
    page = max(1, int(request.args.get("page", 1)))
    per_page = 25
    status_filter = request.args.get("status", "")
    with connect() as conn:
        scope, params = company_scope_sql("t.company_id")
        where = [scope]
        if status_filter:
            where.append("t.status=?")
            params.append(status_filter)
        if not has_permission("task.manage"):
            where.append("(t.assigned_to=? OR t.created_by=?)")
            params.extend([g.v3_user["id"], g.v3_user["id"]])
        where_sql = " AND ".join(where)
        total = conn.execute(f"SELECT COUNT(*) FROM v3_tasks t WHERE {where_sql}", params).fetchone()[0]
        rows = conn.execute(f"SELECT t.*, c.name company_name, u.full_name assigned_name FROM v3_tasks t LEFT JOIN v3_companies c ON c.id=t.company_id LEFT JOIN v3_users u ON u.id=t.assigned_to WHERE {where_sql} ORDER BY CASE WHEN t.status='Tamamlandı' THEN 1 ELSE 0 END, t.due_date LIMIT ? OFFSET ?", params + [per_page, (page-1)*per_page]).fetchall()
        company_rows = conn.execute("SELECT * FROM v3_companies WHERE active=1 ORDER BY name").fetchall() if has_permission("company.all") else conn.execute(f"SELECT * FROM v3_companies WHERE id IN ({','.join('?' for _ in accessible_company_ids())}) ORDER BY name", accessible_company_ids()).fetchall() if accessible_company_ids() else []
        user_rows = conn.execute("SELECT id,full_name FROM v3_users WHERE active=1 ORDER BY full_name").fetchall()
    return render_template("v3/tasks.html", tasks=rows, companies=company_rows, users=user_rows, pager=paginate(page, per_page, total), status_filter=status_filter)


@BP.route("/aging", methods=["GET", "POST"])
def aging():
    if request.method == "POST":
        if not has_permission("aging.manage"):
            abort(403)
        company_id = int(request.form.get("company_id"))
        if company_id not in accessible_company_ids() and not has_permission("company.all"):
            abort(403)
        file = request.files.get("file")
        if not file:
            flash("Dosya seçin.", "danger")
            return redirect(url_for("enterprise_v3.aging"))
        try:
            rows = read_tabular(file)
            source_hash = hashlib.sha256(json.dumps(rows, ensure_ascii=False, default=str).encode()).hexdigest()
            as_of = parse_date(request.form.get("as_of_date")) or date.today().isoformat()
            as_of_date = date.fromisoformat(as_of)
            with connect() as conn:
                if request.form.get("replace") == "1":
                    conn.execute("DELETE FROM v3_aging_items WHERE company_id=? AND audit_id IS ?", (company_id, request.form.get("audit_id") or None))
                count = 0
                for row in rows:
                    doc_date = parse_date(pick(row, "belge tarihi", "fatura tarihi", "document date", "tarih"))
                    due = parse_date(pick(row, "vade tarihi", "due date", "vade")) or doc_date
                    debit = parse_decimal(pick(row, "borç", "debit"))
                    credit = parse_decimal(pick(row, "alacak", "credit"))
                    balance = parse_decimal(pick(row, "bakiye", "balance")) or debit - credit
                    age_days_value = max(0, (as_of_date - date.fromisoformat(due)).days) if due else 0
                    conn.execute("INSERT INTO v3_aging_items(company_id,audit_id,account_code,counterparty_code,counterparty_name,tax_no,document_no,document_date,due_date,debit,credit,balance,age_days,age_bucket,source_hash,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (company_id, request.form.get("audit_id") or None, pick(row, "hesap kodu", "account code"), pick(row, "cari kod", "counterparty code"), pick(row, "cari unvan", "cari adı", "counterparty", "unvan"), pick(row, "vkn", "tckn", "vergi no", "tax no"), pick(row, "belge no", "fatura no", "document no"), doc_date, due, debit, credit, balance, age_days_value, age_bucket(age_days_value), source_hash, now_iso()))
                    count += 1
                conn.commit()
            audit("IMPORT", "aging", source_hash, company_id, new={"rows": count})
            flash(f"{count} yaşlandırma satırı yüklendi.", "success")
        except Exception as exc:
            current_app.logger.exception("Aging import failed")
            flash(f"Dosya okunamadı: {exc}", "danger")
        return redirect(url_for("enterprise_v3.aging", company_id=company_id))
    company_id = request.args.get("company_id", type=int)
    with connect() as conn:
        scope, params = company_scope_sql("company_id")
        if company_id:
            scope += " AND company_id=?"
            params.append(company_id)
        summary = conn.execute(f"SELECT age_bucket, COUNT(*) item_count, SUM(balance) balance FROM v3_aging_items WHERE {scope} GROUP BY age_bucket ORDER BY MIN(age_days)", params).fetchall()
        rows = conn.execute(f"SELECT * FROM v3_aging_items WHERE {scope} ORDER BY age_days DESC LIMIT 500", params).fetchall()
        company_rows = conn.execute("SELECT * FROM v3_companies WHERE active=1 ORDER BY name").fetchall() if has_permission("company.all") else conn.execute(f"SELECT * FROM v3_companies WHERE id IN ({','.join('?' for _ in accessible_company_ids())}) ORDER BY name", accessible_company_ids()).fetchall() if accessible_company_ids() else []
    return render_template("v3/aging.html", rows=rows, summary=summary, companies=company_rows, selected_company=company_id)


def discover_journal_rows(conn: sqlite3.Connection, company_id: int, limit: int = 50000) -> list[dict[str, Any]]:
    excluded = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'v3_%'").fetchall()}
    tables = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'").fetchall() if r[0] not in excluded]
    date_names = {"date", "tarih", "fis_tarihi", "transaction_date", "entry_date"}
    amount_names = {"amount", "tutar", "borc", "borç", "alacak", "debit", "credit"}
    reference_names = {"reference", "referans", "belge_no", "document_no", "fis_no", "fiş_no", "description", "aciklama", "açıklama"}
    best = None
    for table in tables:
        cols = table_columns(conn, table)
        lower = {c.lower(): c for c in cols}
        date_col = next((lower[n] for n in date_names if n in lower), None)
        amount_cols = [lower[n] for n in amount_names if n in lower]
        ref_col = next((lower[n] for n in reference_names if n in lower), None)
        score = (3 if date_col else 0) + len(amount_cols) * 2 + (2 if ref_col else 0)
        if date_col and amount_cols and (best is None or score > best[0]):
            best = (score, table, cols, date_col, amount_cols, ref_col)
    if not best:
        return []
    _, table, cols, date_col, amount_cols, ref_col = best
    id_col = "id" if "id" in cols else "rowid"
    company_col = next((c for c in cols if c.lower() in {"company_id", "firma_id"}), None)
    sql = f'SELECT {id_col} AS _row_id, * FROM "{table}"'
    params: list[Any] = []
    if company_col:
        sql += f' WHERE "{company_col}"=?'
        params.append(company_id)
    sql += f" LIMIT {int(limit)}"
    result = []
    for row in conn.execute(sql, params).fetchall():
        data = dict(row)
        debit = parse_decimal(data.get(next((c for c in amount_cols if c.lower() in {"borc", "borç", "debit"}), "")))
        credit = parse_decimal(data.get(next((c for c in amount_cols if c.lower() in {"alacak", "credit"}), "")))
        amount = parse_decimal(data.get(next((c for c in amount_cols if c.lower() in {"amount", "tutar"}), "")))
        if not amount:
            amount = debit if abs(debit) >= abs(credit) else -credit
        result.append({"table": table, "row_id": data.get("_row_id"), "date": parse_date(data.get(date_col)), "amount": amount, "reference": str(data.get(ref_col, "") if ref_col else "")})
    return result


def match_bank_transactions(conn: sqlite3.Connection, company_id: int, audit_id: Any = None) -> int:
    journal = discover_journal_rows(conn, company_id)
    if not journal:
        return 0
    txns = conn.execute("SELECT * FROM v3_bank_transactions WHERE company_id=? AND audit_id IS ? AND match_status!='Manuel Eşleşti'", (company_id, audit_id)).fetchall()
    matched = 0
    for txn in txns:
        best_score, best_row = 0, None
        txn_date = parse_date(txn["txn_date"])
        for row in journal:
            score = 0
            if abs(abs(txn["amount"]) - abs(row["amount"])) <= 0.01:
                score += 60
            elif abs(abs(txn["amount"]) - abs(row["amount"])) <= 1:
                score += 45
            if txn_date and row["date"]:
                diff = abs((date.fromisoformat(txn_date) - date.fromisoformat(row["date"])).days)
                score += 25 if diff == 0 else 15 if diff <= 2 else 0
            ref = (txn["reference"] or txn["description"] or "").lower()
            jref = (row["reference"] or "").lower()
            if ref and jref and (ref in jref or jref in ref):
                score += 15
            if score > best_score:
                best_score, best_row = score, row
        if best_row and best_score >= 75:
            status = "Eşleşti" if best_score >= 90 else "Muhtemel Eşleşme"
            conn.execute("UPDATE v3_bank_transactions SET match_status=?, match_score=?, journal_table=?, journal_row_id=?, journal_reference=?, journal_date=?, journal_amount=? WHERE id=?", (status, best_score, best_row["table"], str(best_row["row_id"]), best_row["reference"], best_row["date"], best_row["amount"], txn["id"]))
            matched += 1
    conn.commit()
    return matched


@BP.route("/bank-matching", methods=["GET", "POST"])
def bank_matching():
    if request.method == "POST":
        if not has_permission("bank.manage"):
            abort(403)
        company_id = int(request.form.get("company_id"))
        if company_id not in accessible_company_ids() and not has_permission("company.all"):
            abort(403)
        file = request.files.get("file")
        if not file:
            flash("Banka dosyası seçin.", "danger")
            return redirect(url_for("enterprise_v3.bank_matching"))
        try:
            rows = read_tabular(file)
            raw_hash = hashlib.sha256(json.dumps(rows, ensure_ascii=False, default=str).encode()).hexdigest()
            with connect() as conn:
                cur = conn.execute("INSERT INTO v3_bank_imports(company_id,audit_id,filename,sha256,row_count,created_by,created_at) VALUES(?,?,?,?,?,?,?)", (company_id, request.form.get("audit_id") or None, secure_filename(file.filename), raw_hash, len(rows), g.v3_user["id"], now_iso()))
                import_id = cur.lastrowid
                for row in rows:
                    amount = parse_decimal(pick(row, "tutar", "amount", "işlem tutarı", "islem tutari"))
                    incoming = parse_decimal(pick(row, "giriş", "giris", "alacak", "credit"))
                    outgoing = parse_decimal(pick(row, "çıkış", "cikis", "borç", "borc", "debit"))
                    if not amount:
                        amount = incoming - outgoing
                    conn.execute("INSERT INTO v3_bank_transactions(import_id,company_id,audit_id,account_no,txn_date,value_date,reference,description,currency,amount,balance,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)", (import_id, company_id, request.form.get("audit_id") or None, pick(row, "hesap no", "iban", "account"), parse_date(pick(row, "tarih", "işlem tarihi", "transaction date")), parse_date(pick(row, "valör", "valor", "value date")), pick(row, "referans", "reference", "dekont no"), pick(row, "açıklama", "aciklama", "description"), pick(row, "döviz", "doviz", "currency") or "TRY", amount, parse_decimal(pick(row, "bakiye", "balance")) or None, now_iso()))
                conn.commit()
                matched = match_bank_transactions(conn, company_id, request.form.get("audit_id") or None)
            audit("IMPORT", "bank", import_id, company_id, new={"rows": len(rows), "matched": matched})
            flash(f"{len(rows)} hareket yüklendi; {matched} hareket otomatik eşleştirildi.", "success")
        except Exception as exc:
            current_app.logger.exception("Bank import failed")
            flash(f"Banka dosyası okunamadı: {exc}", "danger")
        return redirect(url_for("enterprise_v3.bank_matching", company_id=company_id))
    company_id = request.args.get("company_id", type=int)
    status = request.args.get("status", "")
    page = max(1, request.args.get("page", 1, type=int))
    per_page = 50
    with connect() as conn:
        scope, params = company_scope_sql("b.company_id")
        where = [scope]
        if company_id:
            where.append("b.company_id=?")
            params.append(company_id)
        if status:
            where.append("b.match_status=?")
            params.append(status)
        where_sql = " AND ".join(where)
        total = conn.execute(f"SELECT COUNT(*) FROM v3_bank_transactions b WHERE {where_sql}", params).fetchone()[0]
        rows = conn.execute(f"SELECT b.*, c.name company_name FROM v3_bank_transactions b LEFT JOIN v3_companies c ON c.id=b.company_id WHERE {where_sql} ORDER BY b.txn_date DESC,b.id DESC LIMIT ? OFFSET ?", params + [per_page,(page-1)*per_page]).fetchall()
        summary = conn.execute(f"SELECT match_status,COUNT(*) count,SUM(amount) amount FROM v3_bank_transactions b WHERE {scope} GROUP BY match_status", params[:len(company_scope_sql('b.company_id')[1])]).fetchall()
        company_rows = conn.execute("SELECT * FROM v3_companies WHERE active=1 ORDER BY name").fetchall() if has_permission("company.all") else conn.execute(f"SELECT * FROM v3_companies WHERE id IN ({','.join('?' for _ in accessible_company_ids())}) ORDER BY name", accessible_company_ids()).fetchall() if accessible_company_ids() else []
    return render_template("v3/bank_matching.html", rows=rows, summary=summary, companies=company_rows, selected_company=company_id, selected_status=status, pager=paginate(page, per_page, total))


@BP.route("/reports", methods=["GET", "POST"])
def reports():
    if request.method == "POST":
        if not has_permission("report.manage"):
            abort(403)
        sections = request.form.getlist("sections")
        config = {
            "title": request.form.get("title") or "Sürekli Denetim Yönetim Raporu",
            "confidentiality": request.form.get("confidentiality") or "Gizli ve hizmete özel",
            "sections": sections,
            "footer": request.form.get("footer") or "Sürekli Denetim Merkezi v3.0",
            "signature_name": request.form.get("signature_name"),
            "signature_title": request.form.get("signature_title"),
        }
        with connect() as conn:
            cur = conn.execute("INSERT INTO v3_report_templates(name,company_id,report_type,config_json,active,created_by,created_at,updated_at) VALUES(?,?,?,?,1,?,?,?)", (request.form.get("name", "Yeni Şablon"), request.form.get("company_id") or None, request.form.get("report_type", "Yönetim"), json.dumps(config, ensure_ascii=False), g.v3_user["id"], now_iso(), now_iso()))
            conn.commit()
        audit("CREATE", "report_template", cur.lastrowid, request.form.get("company_id") or None)
        flash("Rapor şablonu kaydedildi.", "success")
        return redirect(url_for("enterprise_v3.reports"))
    with connect() as conn:
        templates = conn.execute("SELECT t.*,c.name company_name FROM v3_report_templates t LEFT JOIN v3_companies c ON c.id=t.company_id WHERE t.active=1 ORDER BY t.updated_at DESC").fetchall()
        companies = conn.execute("SELECT * FROM v3_companies WHERE active=1 ORDER BY name").fetchall()
    return render_template("v3/reports.html", templates=templates, companies=companies)


@BP.route("/reports/<int:template_id>/preview")
def report_preview(template_id: int):
    if not (has_permission("report.view") or has_permission("report.manage")):
        abort(403)
    with connect() as conn:
        template = conn.execute("SELECT * FROM v3_report_templates WHERE id=? AND active=1", (template_id,)).fetchone()
        if not template:
            abort(404)
        config = json.loads(template["config_json"])
        company_id = template["company_id"] or request.args.get("company_id", type=int)
        company = conn.execute("SELECT * FROM v3_companies WHERE id=?", (company_id,)).fetchone() if company_id else None
        task_summary = conn.execute("SELECT status,COUNT(*) count FROM v3_tasks WHERE company_id IS ? GROUP BY status", (company_id,)).fetchall()
        aging_summary = conn.execute("SELECT age_bucket,SUM(balance) balance FROM v3_aging_items WHERE company_id IS ? GROUP BY age_bucket", (company_id,)).fetchall()
        bank_summary = conn.execute("SELECT match_status,COUNT(*) count FROM v3_bank_transactions WHERE company_id IS ? GROUP BY match_status", (company_id,)).fetchall()
    return render_template("v3/report_preview.html", template=template, config=config, company=company, task_summary=task_summary, aging_summary=aging_summary, bank_summary=bank_summary, generated_at=now_iso())


def create_backup_file(password: str | None = None) -> tuple[Path, str, bool]:
    root = Path(current_app.root_path)
    backup_dir = root / "data" / "backups_v3"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    plain = backup_dir / f"surekli_denetim_v3_{stamp}.zip"
    with zipfile.ZipFile(plain, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in (root / "data").rglob("*"):
            if path.is_file() and backup_dir not in path.parents and not path.name.endswith(("-wal", "-shm")):
                zf.write(path, path.relative_to(root))
        manifest = {"version": VERSION, "created_at": now_iso(), "database": str(db_path().relative_to(root) if db_path().is_relative_to(root) else db_path())}
        zf.writestr("V3_BACKUP_MANIFEST.json", json.dumps(manifest, ensure_ascii=False, indent=2))
    encrypted = False
    final_path = plain
    if password:
        try:
            from cryptography.hazmat.primitives import hashes
            from cryptography.hazmat.primitives.ciphers.aead import AESGCM
            from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
            salt = os.urandom(16)
            nonce = os.urandom(12)
            kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=390000)
            key = kdf.derive(password.encode("utf-8"))
            cipher = AESGCM(key).encrypt(nonce, plain.read_bytes(), b"surekli-denetim-v3")
            final_path = plain.with_suffix(".sdbackup")
            final_path.write_bytes(b"SDV3" + salt + nonce + cipher)
            plain.unlink()
            encrypted = True
        except ImportError:
            current_app.logger.warning("cryptography not installed; backup is unencrypted")
    digest = hashlib.sha256(final_path.read_bytes()).hexdigest()
    return final_path, digest, encrypted


@BP.route("/backups", methods=["GET", "POST"])
@permission_required("backup.create")
def backups():
    if request.method == "POST":
        path, digest, encrypted = create_backup_file(request.form.get("password") or None)
        with connect() as conn:
            cur = conn.execute("INSERT INTO v3_backup_jobs(filename,sha256,encrypted,size_bytes,created_by,created_at,verified_at) VALUES(?,?,?,?,?,?,?)", (path.name, digest, 1 if encrypted else 0, path.stat().st_size, g.v3_user["id"], now_iso(), now_iso()))
            conn.commit()
        audit("CREATE", "backup", cur.lastrowid, new={"filename": path.name, "encrypted": encrypted})
        flash("Şifreli yedek oluşturuldu." if encrypted else "Yedek oluşturuldu. Şifreleme paketi yoksa dosya şifresizdir.", "success")
        return redirect(url_for("enterprise_v3.backups"))
    with connect() as conn:
        rows = conn.execute("SELECT b.*,u.full_name created_name FROM v3_backup_jobs b LEFT JOIN v3_users u ON u.id=b.created_by ORDER BY b.created_at DESC").fetchall()
    return render_template("v3/backups.html", backups=rows)


@BP.route("/backups/<int:backup_id>/download")
@permission_required("backup.create")
def backup_download(backup_id: int):
    with connect() as conn:
        row = conn.execute("SELECT * FROM v3_backup_jobs WHERE id=?", (backup_id,)).fetchone()
    if not row:
        abort(404)
    path = Path(current_app.root_path) / "data" / "backups_v3" / row["filename"]
    if not path.exists() or hashlib.sha256(path.read_bytes()).hexdigest() != row["sha256"]:
        abort(409, "Yedek bütünlük kontrolü başarısız")
    audit("DOWNLOAD", "backup", backup_id)
    return send_file(path, as_attachment=True, download_name=path.name)


@BP.route("/audit-log")
@permission_required("log.view")
def audit_log():
    page = max(1, request.args.get("page", 1, type=int))
    per_page = 100
    action = request.args.get("action", "")
    with connect() as conn:
        where, params = [], []
        if action:
            where.append("action=?")
            params.append(action)
        where_sql = " WHERE " + " AND ".join(where) if where else ""
        total = conn.execute("SELECT COUNT(*) FROM v3_audit_log" + where_sql, params).fetchone()[0]
        rows = conn.execute("SELECT * FROM v3_audit_log" + where_sql + " ORDER BY id DESC LIMIT ? OFFSET ?", params + [per_page,(page-1)*per_page]).fetchall()
    return render_template("v3/audit_log.html", rows=rows, selected_action=action, pager=paginate(page, per_page, total))


@BP.route("/notifications/read", methods=["POST"])
def notifications_read():
    with connect() as conn:
        conn.execute("UPDATE v3_notifications SET is_read=1 WHERE user_id=?", (g.v3_user["id"],))
        conn.commit()
    return jsonify({"ok": True})


@BP.errorhandler(403)
def forbidden(error):
    return render_template("v3/error.html", code=403, message="Bu işlem için yetkiniz bulunmuyor."), 403


@BP.errorhandler(404)
def not_found(error):
    return render_template("v3/error.html", code=404, message="Kayıt veya sayfa bulunamadı."), 404
